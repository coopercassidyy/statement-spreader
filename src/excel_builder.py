"""Builds the output workbook: one tab per company (3-statement spread with
linked formulas) plus a cross-referenced Comps Summary tab.

Row layout is fixed across every company sheet so the Comps Summary tab can
point formulas at consistent cell addresses.
"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MILLIONS_FMT = "#,##0,,"
PCT_FMT = "0.0%"
EPS_FMT = "$0.00"

HEADER_FILL = PatternFill("solid", fgColor="1F2937")
SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True)
FORMULA_FONT = Font(bold=True)
OK_FILL = PatternFill("solid", fgColor="C6EFCE")
BAD_FILL = PatternFill("solid", fgColor="FFC7CE")

FIRST_DATA_COL = 2  # column B; column A holds row labels

# Fixed rows. Values are ("label", kind, line_item_key|formula_spec)
# kind: "input" (raw spread value), "formula", "section", "blank"
ROWS = [
    (1, "title", None),
    (2, "blank", None),
    (3, "section", "INCOME STATEMENT ($ in millions)"),
    (4, "input", "revenue"),
    (5, "input", "cost_of_revenue"),
    (6, "formula", "gross_profit"),
    (7, "formula", "gross_margin"),
    (8, "input", "research_and_development"),
    (9, "input", "sga"),
    (10, "input", "operating_income"),
    (11, "formula", "operating_margin"),
    (12, "input", "interest_expense"),
    (13, "input", "income_tax_expense"),
    (14, "input", "net_income"),
    (15, "formula", "net_margin"),
    (16, "input_eps", "eps_diluted"),
    (17, "blank", None),
    (18, "section", "BALANCE SHEET ($ in millions)"),
    (19, "input", "cash"),
    (20, "input", "current_assets"),
    (21, "input", "ppe_net"),
    (22, "input", "total_assets"),
    (23, "input", "current_liabilities"),
    (24, "input", "long_term_debt"),
    (25, "input", "total_liabilities"),
    (26, "input", "stockholders_equity"),
    (27, "formula", "balance_check"),
    (28, "blank", None),
    (29, "section", "CASH FLOW STATEMENT ($ in millions)"),
    (30, "input", "cfo"),
    (31, "input", "cfi"),
    (32, "input", "cff"),
    (33, "formula", "net_change_in_cash"),
    (34, "input", "capex"),
    (35, "input", "depreciation_amortization"),
]

ROW_LABELS = {
    "revenue": "Revenue", "cost_of_revenue": "Cost of Revenue",
    "gross_profit": "Gross Profit", "gross_margin": "  Gross Margin %",
    "research_and_development": "Research & Development", "sga": "SG&A",
    "operating_income": "Operating Income", "operating_margin": "  Operating Margin %",
    "interest_expense": "Interest Expense", "income_tax_expense": "Income Tax Expense",
    "net_income": "Net Income", "net_margin": "  Net Margin %", "eps_diluted": "Diluted EPS",
    "cash": "Cash & Equivalents", "current_assets": "Total Current Assets",
    "ppe_net": "PP&E, Net", "total_assets": "Total Assets",
    "current_liabilities": "Total Current Liabilities", "long_term_debt": "Long-Term Debt",
    "total_liabilities": "Total Liabilities", "stockholders_equity": "Stockholders' Equity",
    "balance_check": "Balance Check (Assets − (Liab + Equity))",
    "cfo": "Cash from Operations", "cfi": "Cash from Investing", "cff": "Cash from Financing",
    "net_change_in_cash": "Net Change in Cash", "capex": "Capital Expenditures",
    "depreciation_amortization": "Depreciation & Amortization",
}

ROW_BY_KEY = {spec: row for row, kind, spec in ROWS if kind in ("input", "input_eps", "formula")}


def _safe_sheet_name(ticker: str) -> str:
    return ticker[:31]


def _write_company_sheet(wb: Workbook, ticker: str, entity_name: str, spread: dict, validation: dict):
    ws = wb.create_sheet(_safe_sheet_name(ticker))
    years = sorted(spread.keys())
    cols = {fy: FIRST_DATA_COL + i for i, fy in enumerate(years)}

    ws.cell(row=1, column=1, value=f"{ticker} — {entity_name}").font = Font(bold=True, size=14)
    for fy, col in cols.items():
        c = ws.cell(row=1, column=col, value=f"FY{fy}")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")

    for row, kind, spec in ROWS:
        if kind == "title":
            continue
        if kind == "section":
            ws.cell(row=row, column=1, value=spec).font = SECTION_FONT
            for col in range(1, FIRST_DATA_COL + len(years)):
                ws.cell(row=row, column=col).fill = SECTION_FILL
            continue
        if kind == "blank":
            continue

        label = ROW_LABELS[spec]
        ws.cell(row=row, column=1, value=label)

        for fy, col in cols.items():
            year_data = spread[fy]
            col_letter = get_column_letter(col)

            if kind == "input":
                item = year_data.get(spec, {})
                value = item.get("value")
                cell = ws.cell(row=row, column=col, value=value / 1_000_000 if value is not None else None)
                if value is not None:
                    cell.number_format = "#,##0"
                if item.get("derived"):
                    cell.font = Font(italic=True)

            elif kind == "input_eps":
                item = year_data.get(spec, {})
                value = item.get("value")
                cell = ws.cell(row=row, column=col, value=value)
                if value is not None:
                    cell.number_format = EPS_FMT

            elif kind == "formula":
                cell = ws.cell(row=row, column=col)
                cell.font = FORMULA_FONT
                if spec == "gross_profit":
                    cell.value = f"={col_letter}4-{col_letter}5"
                    cell.number_format = "#,##0"
                elif spec == "gross_margin":
                    cell.value = f"=IFERROR({col_letter}6/{col_letter}4,\"\")"
                    cell.number_format = PCT_FMT
                elif spec == "operating_margin":
                    cell.value = f"=IFERROR({col_letter}10/{col_letter}4,\"\")"
                    cell.number_format = PCT_FMT
                elif spec == "net_margin":
                    cell.value = f"=IFERROR({col_letter}14/{col_letter}4,\"\")"
                    cell.number_format = PCT_FMT
                elif spec == "net_change_in_cash":
                    cell.value = f"={col_letter}30+{col_letter}31+{col_letter}32"
                    cell.number_format = "#,##0"
                elif spec == "balance_check":
                    formula = (
                        f'=IF(OR({col_letter}22="",{col_letter}25="",{col_letter}26=""),"n/a",'
                        f'IF(ABS({col_letter}22-({col_letter}25+{col_letter}26))<1,"OK","MISMATCH"))'
                    )
                    cell.value = formula
                    check = validation.get(fy, {}).get("balance_sheet_ties", {})
                    if check.get("ok") is True:
                        cell.fill = OK_FILL
                    elif check.get("ok") is False:
                        cell.fill = BAD_FILL

    if any(item["derived"] for yd in spread.values() for k, item in yd.items() if k == "total_liabilities"):
        ws.cell(row=36, column=1, value="* Total Liabilities derived as Total Assets − Stockholders' Equity "
                                          "where not directly reported.").font = Font(italic=True, size=9)

    ws.column_dimensions["A"].width = 32
    for col in cols.values():
        ws.column_dimensions[get_column_letter(col)].width = 14

    return years


def _write_comps_sheet(wb: Workbook, company_years: dict):
    ws = wb.create_sheet("Comps Summary", 0)
    headers = [
        "Ticker", "Company", "Latest FY", "Revenue ($M)", "Revenue YoY %",
        "Gross Margin %", "Operating Margin %", "Net Margin %", "Total Assets ($M)", "ROE %",
    ]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL

    for r, (ticker, info) in enumerate(company_years.items(), start=2):
        years = info["years"]
        entity_name = info["entity_name"]
        sheet = _safe_sheet_name(ticker)
        latest_col = get_column_letter(FIRST_DATA_COL + len(years) - 1)
        prior_col = get_column_letter(FIRST_DATA_COL + len(years) - 2) if len(years) > 1 else None

        ws.cell(row=r, column=1, value=ticker)
        ws.cell(row=r, column=2, value=entity_name)
        ws.cell(row=r, column=3, value=f"='{sheet}'!{latest_col}1")
        ws.cell(row=r, column=4, value=f"='{sheet}'!{latest_col}4").number_format = "#,##0"
        if prior_col:
            ws.cell(row=r, column=5,
                     value=f"=IFERROR(('{sheet}'!{latest_col}4-'{sheet}'!{prior_col}4)/'{sheet}'!{prior_col}4,\"\")"
                     ).number_format = PCT_FMT
        ws.cell(row=r, column=6, value=f"='{sheet}'!{latest_col}7").number_format = PCT_FMT
        ws.cell(row=r, column=7, value=f"='{sheet}'!{latest_col}11").number_format = PCT_FMT
        ws.cell(row=r, column=8, value=f"='{sheet}'!{latest_col}15").number_format = PCT_FMT
        ws.cell(row=r, column=9, value=f"='{sheet}'!{latest_col}22").number_format = "#,##0"
        ws.cell(row=r, column=10,
                 value=f"=IFERROR('{sheet}'!{latest_col}14/'{sheet}'!{latest_col}26,\"\")"
                 ).number_format = PCT_FMT

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 24
    for col in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16


def build_workbook(companies: dict) -> Workbook:
    """companies: {ticker: {"entity_name": str, "spread": dict, "validation": dict}}"""
    wb = Workbook()
    wb.remove(wb.active)

    company_years = {}
    for ticker, data in companies.items():
        years = _write_company_sheet(wb, ticker, data["entity_name"], data["spread"], data["validation"])
        company_years[ticker] = {"years": years, "entity_name": data["entity_name"]}

    _write_comps_sheet(wb, company_years)
    return wb
